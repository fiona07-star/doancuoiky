import xlsxwriter as xr
from openpyxl import load_workbook
import os

from Spydecat_K24406H.models.Nganh_hoc import Nganh_hoc


class ExportTool:
    def export_bang_diem_to_excel(self, filename, danh_sach_nganh):
        try:
            workbook = xr.Workbook(filename)
            worksheet = workbook.add_worksheet()

            # Thiết lập độ rộng cột
            worksheet.set_column('A:A', 25)  # Tên ngành
            worksheet.set_column('B:B', 30)  # Chuyên ngành
            worksheet.set_column('C:C', 15)  # Mã tuyển sinh
            worksheet.set_column('D:D', 10)  # Chỉ tiêu
            worksheet.set_column('E:E', 10)  # PT1A
            worksheet.set_column('F:F', 10)  # PT1B
            worksheet.set_column('G:G', 10)  # PT2
            worksheet.set_column('H:H', 10)  # PT3

            # Format chữ đậm cho tiêu đề
            bold = workbook.add_format({'bold': True})

            # Thêm tiêu đề cột
            headers = ["Tên ngành", "Chuyên ngành", "Mã tuyển sinh", "Chỉ tiêu", "PT1A", "PT1B", "PT2", "PT3"]
            for col, header in enumerate(headers):
                worksheet.write(0, col, header, bold)

            # Ghi dữ liệu từ danh sách ngành học
            for i, nh in enumerate(danh_sach_nganh, start=1):
                worksheet.write(i, 0, nh.ten_nganh)
                worksheet.write(i, 1, nh.chuyen_nganh)
                worksheet.write(i, 2, nh.ma_tuyen_sinh)
                worksheet.write(i, 3, nh.chi_tieu)
                worksheet.write(i, 4, nh.phuong_thuc_1a)
                worksheet.write(i, 5, nh.phuong_thuc_1b)
                worksheet.write(i, 6, nh.phuong_thuc_2)
                worksheet.write(i, 7, nh.phuong_thuc_3)

            workbook.close()
            print("Xuất file Excel thành công!")

        except PermissionError:
            print("LỖI: Không thể ghi vào file. Hãy đóng file Excel nếu đang mở!")

    def import_bang_diem_from_excel(self, filename):
        try:
            # Kiểm tra file có tồn tại không
            if not os.path.exists(filename):
                print(f"LỖI: File '{filename}' không tồn tại!")
                return []

            wb = load_workbook(filename, data_only=True)
            ws = wb.active
            is_header = True
            danh_sach_nganh = []

            for row in ws.iter_rows(values_only=True):
                if is_header:
                    is_header = False
                    continue  # Bỏ qua dòng tiêu đề

                # Kiểm tra xem dòng có đủ dữ liệu không
                if all(cell is not None for cell in row):
                    try:
                        ten_nganh, chuyen_nganh, ma_tuyen_sinh, chi_tieu, pt1a, pt1b, pt2, pt3 = row
                        ngh = Nganh_hoc(ten_nganh, chuyen_nganh, ma_tuyen_sinh, chi_tieu, pt1a, pt1b, pt2, pt3)
                        danh_sach_nganh.append(ngh)
                    except ValueError:
                        print(f"Cảnh báo: Lỗi dữ liệu tại dòng {row}")

            wb.close()
            print("Import dữ liệu từ file Excel thành công!")
            return danh_sach_nganh

        except PermissionError:
            print("LỖI: Không thể mở file. Hãy đóng file Excel nếu đang mở!")

    def export_danhsach_khach_to_excel(self, filename, danhsach_khach):
        try:
            workbook = xr.Workbook(filename)
            worksheet = workbook.add_worksheet()

            # Thiết lập độ rộng cột
            worksheet.set_column('A:A', 15)  # Mã người dùng
            worksheet.set_column('B:B', 25)  # Họ tên
            worksheet.set_column('C:C', 30)  # Email
            worksheet.set_column('D:D', 15)  # Số điện thoại
            worksheet.set_column('E:E', 10)  # Giới tính
            worksheet.set_column('F:F', 10)  # Năm sinh

            # Format tiêu đề (in đậm)
            bold = workbook.add_format({'bold': True})

            # Thêm tiêu đề cột
            headers = ["Mã người dùng", "Họ tên", "Email", "Số điện thoại", "Giới tính", "Năm sinh"]
            for col, header in enumerate(headers):
                worksheet.write(0, col, header, bold)

            # Ghi dữ liệu danh sách khách vào file Excel
            for i, kh in enumerate(danhsach_khach, start=1):
                worksheet.write(i, 0, kh.ma_nguoi_dung)
                worksheet.write(i, 1, kh.ten)
                worksheet.write(i, 2, kh.email)
                worksheet.write(i, 3, kh.sdt)
                worksheet.write(i, 4, "Nam" if kh.gioi_tinh else "Nữ")  # Chuyển đổi True/False thành Nam/Nữ
                worksheet.write(i, 5, kh.nam_sinh)

            workbook.close()
            print("Xuất danh sách khách thành công!")

        except PermissionError:
            print("LỖI: Không thể ghi vào file. Hãy đóng file Excel nếu đang mở!")
